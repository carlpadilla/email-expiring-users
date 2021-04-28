#! python3
# account_expiring_reminder.py - Sends emails based on expiration date in spreadsheet.
import smtplib
from datetime import timedelta, datetime

import openpyxl

# 15 days until password expires
endDate = datetime.today() + timedelta(days=15)


# spreadsheet should have three columns name, email, expiration date
wb = openpyxl.load_workbook('expiring_accounts.xlsx')

sheet = wb['Sheet1']

lastCol = sheet.max_column

expiryMonth = sheet.cell(row=3, column=lastCol).value

# Check each member's expiry status.
expiringUsers = {}
for r in range(3, sheet.max_row + 1):
    expiry = sheet.cell(row=r, column=lastCol).value
    # print(expiry)

    if expiry < endDate:
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        expiringUsers[name] = email
# uncomment to view users in expiringUsers Obj
print(expiringUsers)

#  Log in to email account.
# Enter smtp address and port number
smtpObj = smtplib.SMTP('smtp.office365.com', 587)

# establishing a connection to the server
smtpObj.ehlo()

# enables TLS encryption for your connection on port 587
smtpObj.starttls()


#
smtpObj.login('{ENTER SENDING EMAIL ADDRESS HERE}',
              input('Enter email password: '))

''' sendmail() method requires three arguments. from address, recipient's, and email body.
The start of the email body string must begin with 'Subject: \n' for the
subject line of the email. The '\n' newline character separates the subject
line from the main body of the email.
'''


# TODO: Send out reminder emails.
for name, email in expiringUsers.items():
    message = f"""\
Subject: {name} Domain Account Expiring soon.

Dear {name},

Your Domain password will expire on {expiry}. Please change it as soon as possible.
To Change your password, follow the method below:
1. On your Windows computer
    a. If you're not in the office, connect to VPN.
    b. Log into your computer as usual and make sure you are connected to the internet (VPN if remote).
    c. Press Ctrl-Alt-Del and click on 'Change Password'.
    d. Fill in your old password and set a new password.
    e. Press OK to return to your desktop
    
    Please make sure to do this before the Expiration date or you will be locked out.
    
    """
    print(f'Sending email to {email}...')
    sendmailStatus = smtpObj.sendmail(
        '{ENTER SENDING EMAIL ADDRESS HERE}', email, message)
    if sendmailStatus != {}:
        print(f'There was a problem sending email to {name}: {email}')
        smtpObj.quit()
